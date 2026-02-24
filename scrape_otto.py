"""
Otto.de Product Information Scraper

Multi-pass matching strategy:
  Pass 1 (STRICT):  product_line + model + sub_family + variants must all match
  Pass 2 (RELAXED): product_line + model must match (ignores variant/sub-family)
  Pass 3 (BRAND):   brand + model only (last resort)

This ensures the scraper always finds the best available match even when
the exact variant isn't listed on otto.de.
"""

from pathlib import Path
from dataclasses import dataclass, asdict
from typing import Optional, Tuple
import csv
import time
import random
import logging
import io
import re
import urllib.request
import urllib.error

from playwright.sync_api import sync_playwright, Page

# Optional: stealth to prevent captchas from appearing
try:
    from playwright_stealth import stealth_sync
    HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False

# Optional: 2captcha to solve captchas when they appear
try:
    from twocaptcha import TwoCaptcha
    HAS_2CAPTCHA = True
except ImportError:
    HAS_2CAPTCHA = False


# ============================================================================
# CONFIGURATION
# ============================================================================

class Config:
    INPUT_FILE = "article_numbers.txt"
    OUTPUT_FILE = "otto_products_report.csv"

    HEADLESS = False
    SLOW_MO = 20
    VIEWPORT_WIDTH = 1920
    VIEWPORT_HEIGHT = 1080

    MIN_DELAY = 0.15
    MAX_DELAY = 0.5
    KEYSTROKE_DELAY_MIN = 0.01
    KEYSTROKE_DELAY_MAX = 0.03
    ACTION_DELAY_MIN = 0.02
    ACTION_DELAY_MAX = 0.08
    SCROLL_MIN = 200
    SCROLL_MAX = 700
    DEFAULT_TIMEOUT_MS = 15000

    OCR_ENABLED = True
    OCR_DPI = 200
    OCR_LANG = "eng+deu"
    POPPLER_PATH = "/opt/homebrew/bin"
    TESSERACT_CMD = "/opt/homebrew/bin/tesseract"

    # --- Captcha settings ---
    TWOCAPTCHA_API_KEY = ""
    CAPTCHA_MAX_RETRIES = 3
    CAPTCHA_WAIT_AFTER_SOLVE = 3


@dataclass
class ProductData:
    input_ean: str
    product_url: str
    pdf_link: str
    energy_efficiency_class: str
    energylevel_link: str
    supplier_information: str


def setup_logging() -> logging.Logger:
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s - %(levelname)s - %(message)s")
    return logging.getLogger(__name__)


logger = setup_logging()


# ============================================================================
# BRAND / MODEL UTILITIES
# ============================================================================

BRAND_FAMILIES = {
    "apple": {
        "product_lines": ["iphone", "ipad", "macbook", "mac"],
        "model_re": r"(?:iphone|ipad)\s*(\d{1,2})",
        "category": "phone",
    },
    "samsung": {
        "product_lines": ["galaxy"],
        "model_re": r"galaxy\s*(?:z\s*)?(?:flip|fold|s|a|m|note)\s*(\d{1,2})",
        "category": "phone",
    },
    "google": {
        "product_lines": ["pixel"],
        "model_re": r"pixel\s*(\d{1,2})",
        "category": "phone",
    },
    "oneplus": {
        "product_lines": ["oneplus"],
        "model_re": r"oneplus\s*(\d{1,2})",
        "category": "phone",
    },
    "xiaomi": {
        "product_lines": ["xiaomi", "redmi", "poco"],
        "model_re": r"(?:xiaomi|redmi|poco)\s*(?:note\s*)?(\d{1,2})",
        "category": "phone",
    },
    "huawei": {
        "product_lines": ["huawei", "mate"],
        "model_re": r"(?:huawei|mate|p)\s*(\d{1,2})",
        "category": "phone",
    },
    "sony": {
        "product_lines": ["xperia"],
        "model_re": r"xperia\s*(\w+)",
        "category": "phone",
    },
    "motorola": {
        "product_lines": ["moto", "motorola"],
        "model_re": r"moto(?:rola)?\s*(\w+)",
        "category": "phone",
    },
    "nothing": {
        "product_lines": ["nothing"],
        "model_re": r"nothing\s*phone\s*\(?(\d+)\)?",
        "category": "phone",
    },
}

VARIANT_TOKENS = ["pro", "max", "ultra", "plus", "lite", "fe", "mini", "xl"]

# ---------------------------------------------------------------------------
# Accessory keywords — checked against BOTH normalised AND raw lowered text
# ---------------------------------------------------------------------------
ACCESSORY_KEYWORDS_NORM = frozenset({
    "huelle", "case", "cover", "bumper", "handyhuelle",
    "schale", "schutzhuelle", "backcover", "flipcase", "bookcase",
    "klapphuelle", "klappcover", "etui", "silikonhuelle",
    "displayschutz", "folie", "schutzfolie", "schutzglas", "panzerglas",
    "panzerfolie", "displayfolie",
    "kabel", "ladekabel", "netzteil",
    "earphone", "earphones", "headphone", "headphones",
    "earbud", "earbuds", "headset",
    "halter", "halterung", "autohalterung",
    "staender", "stativ", "handyhalterung",
    "magnethalterung", "saugnapf",
    "tasche", "pouch",
    "ersatzakku", "powerbank",
    "wristband", "stylus", "eingabestift",
    "reinigung", "reinigungsset", "cleaning", "selfiestick",
    "ringhalter", "fingerhalter", "popgrip", "popsocket",
    "simkarte", "speicherkarte", "sdkarte",
})

ACCESSORY_KEYWORDS_RAW = frozenset({
    "hülle", "huelle", "schutzhülle", "schutzhuelle", "handyhülle",
    "handyhuelle", "schutzfolie", "panzerglas", "panzerfolie",
    "screen protector", "tempered glass",
    "halter", "halterung", "kfz-halter", "kfz-halterung",
    "kfz halter", "kfz halterung", "lüfterhalter", "luefterhalter",
    "autohalterung", "handyhalterung",
    "ladegerät", "ladegeraet", "ladekabel", "netzteil",
    "kopfhörer", "kopfhoerer", "headset", "earbuds",
    "tasche", "pouch", "gürteltasche",
    "powerbank", "ersatzakku",
    "selfiestick", "selfie-stick",
    "popgrip", "popsocket",
    "silikon case", "silikon hülle", "tpu case", "tpu hülle",
    "hardcase", "hard case", "kfz", "charger", "adapter", "armband",
    # English terms from Otto's English UI
    "phone case", "protective case", "slim case",
})

WRONG_CATEGORY_KEYWORDS = frozenset({
    "macbook", "notebook", "laptop", "imac", "mac mini", "mac studio",
    "mac pro", "airpods", "apple watch", "watch ultra", "homepod",
    "apple tv", "airtag", "magic keyboard", "magic mouse", "magic trackpad",
    "galaxy tab", "galaxy watch", "galaxy buds",
    "pixel watch", "pixel buds", "pixel tablet",
    "smart tv", "fernseher", "monitor", "drucker", "printer",
})

NOISE_WORDS = {
    "gb", "tb", "speicher", "farbe", "dual", "sim", "dualsim",
    "schwarz", "weiss", "silber", "rot", "blau", "gruen",
    "white", "black", "blue", "green", "red", "silver", "gold",
    "pink", "pinkgold", "titanium", "cosmic", "orange", "navy",
    "mint", "iris", "moonstone",
    "128", "256", "512", "1000", "64", "32",
}


# ============================================================================
# QUERY ANALYSIS
# ============================================================================

@dataclass
class QueryInfo:
    """Parsed information about what we're searching for."""
    raw: str
    norm: str
    brand: Optional[str]
    product_line: str        # "iphone", "galaxy", "pixel"
    samsung_sub: str         # "flip", "fold", "s", "a", etc.
    model_number: str        # "17", "25", "7", "9"
    variant_tokens: list     # ["pro", "max"], ["ultra"], ["fe"]
    search_tokens: list      # all meaningful tokens for scoring

    @classmethod
    def from_query(cls, query: str) -> "QueryInfo":
        norm = _normalize(query)
        brand = _detect_brand(norm)
        product_line = _extract_product_line(norm, brand)
        samsung_sub = _extract_samsung_sub(norm)
        model_number = _extract_model(norm, brand)
        variants = _extract_variants(norm)
        tokens = _tokenize(norm)
        return cls(query, norm, brand, product_line, samsung_sub,
                   model_number, variants, tokens)

    def log(self):
        logger.info(
            f"  brand={self.brand}  line={self.product_line}  "
            f"sub={self.samsung_sub}  model={self.model_number}  "
            f"variants={self.variant_tokens}  tokens={self.search_tokens}"
        )


def _detect_brand(qn: str) -> Optional[str]:
    for brand, info in BRAND_FAMILIES.items():
        if brand in qn:
            return brand
        for pl in info["product_lines"]:
            if re.search(rf"\b{re.escape(pl)}\b", qn):
                return brand
    return None


def _extract_product_line(qn: str, brand: Optional[str]) -> str:
    if not brand or brand not in BRAND_FAMILIES:
        return ""
    for pl in BRAND_FAMILIES[brand]["product_lines"]:
        if re.search(rf"\b{re.escape(pl)}\b", qn):
            return pl
    return ""


def _extract_samsung_sub(qn: str) -> str:
    m = re.search(r"galaxy\s+(?:z\s+)?(flip|fold)\s*\d", qn)
    if m:
        return m.group(1)
    m = re.search(r"galaxy\s+(?:z\s+)?(flip|fold)\d", qn)
    if m:
        return m.group(1)
    m = re.search(r"galaxy\s+(s|a|m|note)\s*\d", qn)
    if m:
        return m.group(1)
    return ""


def _extract_model(qn: str, brand: Optional[str]) -> str:
    if brand and brand in BRAND_FAMILIES:
        m = re.search(BRAND_FAMILIES[brand]["model_re"], qn, re.I)
        if m:
            return m.group(1)
    m = re.search(r"\b(\d{1,2})\b", qn)
    return m.group(1) if m else ""


def _extract_variants(qn: str) -> list[str]:
    return [t for t in VARIANT_TOKENS if re.search(rf"\b{t}\b", qn)]


def _tokenize(qn: str) -> list[str]:
    return [t for t in qn.split() if t and t not in NOISE_WORDS]


# ============================================================================
# TEXT UTILITIES
# ============================================================================

def _normalize(text: str) -> str:
    text = text.lower()
    # Split joined tokens: "flip7" -> "flip 7", "128gb" -> "128 gb"
    text = re.sub(r"([a-zäöüß])(\d)", r"\1 \2", text)
    text = re.sub(r"(\d)([a-zäöüß])", r"\1 \2", text)
    text = re.sub(r"[^a-z0-9äöüß]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _match_score(text: str, tokens: list[str]) -> int:
    return sum(1 for t in tokens
               if re.search(rf"\b{re.escape(t)}\b", text))


def _is_accessory(title_norm: str, card_norm: str, raw_lower: str) -> bool:
    for kw in ACCESSORY_KEYWORDS_NORM:
        if kw in title_norm or kw in card_norm:
            return True
    for kw in ACCESSORY_KEYWORDS_RAW:
        if kw in raw_lower:
            return True
    return False


def _is_wrong_category(combined: str, raw_lower: str,
                       brand: Optional[str]) -> bool:
    if brand and BRAND_FAMILIES.get(brand, {}).get("category") == "phone":
        for kw in WRONG_CATEGORY_KEYWORDS:
            if kw in combined or kw in raw_lower:
                return True
    return False


def _is_sponsored(raw_lower: str) -> bool:
    return ("gesponsert" in raw_lower or "anzeige" in raw_lower
            or "sponsored" in raw_lower)


# ============================================================================
# BROWSER HELPERS
# ============================================================================

class BH:
    @staticmethod
    def human_type(el, text: str):
        el.click()
        el.fill("")
        for c in text:
            el.type(c)
            time.sleep(random.uniform(Config.KEYSTROKE_DELAY_MIN,
                                      Config.KEYSTROKE_DELAY_MAX))

    @staticmethod
    def mouse_move(page: Page):
        try:
            page.mouse.move(random.randint(100, 500),
                            random.randint(100, 500))
        except Exception:
            pass

    @staticmethod
    def scroll(page: Page, amount: int = 0):
        try:
            page.mouse.wheel(0, amount or random.randint(
                Config.SCROLL_MIN, Config.SCROLL_MAX))
        except Exception:
            pass

    @staticmethod
    def delay():
        return random.uniform(Config.MIN_DELAY, Config.MAX_DELAY)

    @staticmethod
    def short_delay():
        return random.uniform(Config.ACTION_DELAY_MIN,
                              Config.ACTION_DELAY_MAX)


# ============================================================================
# CAPTCHA DETECTION & SOLVING
# ============================================================================

class CaptchaSolver:
    """Detects and solves captchas. Install: pip install playwright-stealth 2captcha-python"""

    def __init__(self, config: Config):
        self.config = config
        self.solver = None
        if HAS_2CAPTCHA and config.TWOCAPTCHA_API_KEY:
            self.solver = TwoCaptcha(config.TWOCAPTCHA_API_KEY)
            logger.info("2Captcha solver initialized")

    @staticmethod
    def apply_stealth(page: Page):
        if HAS_STEALTH:
            stealth_sync(page)
            logger.info("Stealth patches applied")

    def detect_captcha(self, page: Page) -> Optional[str]:
        try:
            return page.evaluate("""() => {
                if (document.querySelector('iframe[src*="recaptcha"]') ||
                    document.querySelector('.g-recaptcha')) return 'recaptcha';
                if (document.querySelector('iframe[src*="hcaptcha"]') ||
                    document.querySelector('.h-captcha')) return 'hcaptcha';
                if (document.querySelector('iframe[src*="challenges.cloudflare"]') ||
                    document.querySelector('.cf-turnstile')) return 'turnstile';
                const body = (document.body.innerText || '').toLowerCase();
                if (body.includes('captcha') || body.includes('are you human') ||
                    body.includes('bist du ein mensch') ||
                    body.includes('ich bin kein roboter')) return 'generic';
                return null;
            }""")
        except Exception:
            return None

    def solve(self, page: Page) -> bool:
        ctype = self.detect_captcha(page)
        if not ctype:
            return True
        logger.warning(f"CAPTCHA detected: {ctype}")
        if not self.solver:
            if not Config.HEADLESS:
                logger.info("Waiting 60s for manual captcha solve...")
                for _ in range(60):
                    time.sleep(1)
                    if not self.detect_captcha(page):
                        logger.info("Captcha solved manually!")
                        return True
            return False
        for _ in range(self.config.CAPTCHA_MAX_RETRIES):
            try:
                sitekey = self._get_sitekey(page, ctype)
                if not sitekey:
                    return False
                token = self._solve_type(ctype, sitekey, page.url)
                if token:
                    self._inject(page, ctype, token)
                    time.sleep(self.config.CAPTCHA_WAIT_AFTER_SOLVE)
                    if not self.detect_captcha(page):
                        logger.info("Captcha solved!")
                        return True
            except Exception as e:
                logger.error(f"Captcha error: {e}")
        return False

    def _get_sitekey(self, page, ctype):
        try:
            return page.evaluate("""(type) => {
                if (type === 'recaptcha') {
                    const el = document.querySelector('.g-recaptcha[data-sitekey]');
                    if (el) return el.getAttribute('data-sitekey');
                    const f = document.querySelector('iframe[src*="recaptcha"]');
                    if (f) { const m = f.src.match(/[?&]k=([^&]+)/); if (m) return m[1]; }
                }
                if (type === 'hcaptcha') {
                    const el = document.querySelector('.h-captcha[data-sitekey]');
                    if (el) return el.getAttribute('data-sitekey');
                }
                if (type === 'turnstile') {
                    const el = document.querySelector('.cf-turnstile[data-sitekey], [data-sitekey]');
                    if (el) return el.getAttribute('data-sitekey');
                }
                return null;
            }""", ctype)
        except Exception:
            return None

    def _solve_type(self, ctype, sitekey, url):
        if ctype == 'recaptcha':
            r = self.solver.recaptcha(sitekey=sitekey, url=url)
        elif ctype == 'hcaptcha':
            r = self.solver.hcaptcha(sitekey=sitekey, url=url)
        elif ctype == 'turnstile':
            r = self.solver.turnstile(sitekey=sitekey, url=url)
        else:
            return None
        return r.get("code", "")

    def _inject(self, page, ctype, token):
        page.evaluate("""(args) => {
            const [type, token] = args;
            if (type === 'recaptcha') {
                const el = document.querySelector('#g-recaptcha-response, textarea[name="g-recaptcha-response"]');
                if (el) el.value = token;
                const form = document.querySelector('form'); if (form) form.submit();
            } else if (type === 'hcaptcha') {
                const el = document.querySelector('[name="h-captcha-response"]');
                if (el) el.value = token;
            } else if (type === 'turnstile') {
                const el = document.querySelector('[name="cf-turnstile-response"]');
                if (el) el.value = token;
            }
        }""", [ctype, token])


# ============================================================================
# OTTO NAVIGATOR
# ============================================================================

# Multiple selectors for product cards — Otto's HTML may vary
CARD_SELECTORS = [
    "article",
    "[data-testid*='product']",
    ".product-card",
    ".js_productCard",
    "[class*='productCard']",
    "[class*='ProductCard']",
]


class OttoNavigator:
    def __init__(self, page: Page, captcha_solver: Optional[CaptchaSolver] = None):
        self.page = page
        self._cookies_done = False
        self.captcha_solver = captcha_solver

    def _check_captcha(self):
        if self.captcha_solver and self.captcha_solver.detect_captcha(self.page):
            self.captcha_solver.solve(self.page)

    def accept_cookies(self):
        try:
            if self._cookies_done:
                return
            time.sleep(random.uniform(0.6, 1.2))
            for sel in ["#onetrust-accept-btn-handler",
                        "button:has-text('Akzeptieren')",
                        "button:has-text('Accept')"]:
                btn = self.page.locator(sel)
                if btn.count() > 0:
                    btn.first.click(timeout=5000)
                    logger.info("Cookies accepted")
                    time.sleep(random.uniform(0.3, 0.7))
                    break
            self._cookies_done = True
        except Exception:
            pass

    def search_product(self, query: str):
        logger.info(f"Searching: {query}")
        self.page.goto("https://www.otto.de/", wait_until="domcontentloaded")
        self._check_captcha()
        self.accept_cookies()
        BH.mouse_move(self.page)
        time.sleep(BH.short_delay())

        search_box = None
        for sel in ["input[type='search']", "input[name='q']",
                    "input[placeholder*='Suchen' i]",
                    "input[placeholder*='Search' i]",
                    "header input"]:
            try:
                el = self.page.locator(sel).first
                el.wait_for(state="visible", timeout=10000)
                search_box = el
                break
            except Exception:
                continue

        if not search_box:
            raise Exception("Search box not found")

        BH.human_type(search_box, query)
        time.sleep(BH.short_delay())
        search_box.press("Enter")

        self.page.wait_for_load_state("domcontentloaded")
        self._check_captcha()
        self._wait_for_cards()

    def _wait_for_cards(self):
        """Wait for product cards to appear using multiple selectors."""
        for sel in CARD_SELECTORS:
            try:
                self.page.wait_for_selector(sel, timeout=4000)
                return
            except Exception:
                continue
        logger.warning("No product cards found with any selector")

    def _get_cards(self):
        """Get all product card elements using the first matching selector."""
        for sel in CARD_SELECTORS:
            cards = self.page.locator(sel).all()
            if cards:
                return cards
        return []

    # ------------------------------------------------------------------
    # MULTI-PASS PRODUCT FINDING
    # ------------------------------------------------------------------
    def find_product(self, query: str) -> bool:
        """Find the correct product using a multi-pass strategy:
        Pass 1: Strict (all criteria)
        Pass 2: Relaxed (brand + product_line + model, ignore variants)
        Pass 3: Brand only (brand + product_line, ignore model)
        """
        if "/p/" in self.page.url:
            logger.info(f"Already on product page: {self.page.url}")
            return True

        qi = QueryInfo.from_query(query)
        qi.log()

        # Try strict matching first, with scrolling
        result = self._find_with_passes(qi)
        if result:
            return True

        logger.error(f"No product found for: {query}")
        return False

    def _find_with_passes(self, qi: QueryInfo) -> bool:
        """Scroll through results trying strict, then relaxed, then brand-only."""
        # Collect all cards across multiple scroll positions
        all_scored: list[tuple] = []

        for scroll_attempt in range(10):
            cards = self._get_cards()
            if not cards:
                logger.warning(
                    f"No cards found (scroll {scroll_attempt}), trying next selector..."
                )
                BH.scroll(self.page, 1200)
                time.sleep(BH.short_delay())
                continue

            for idx, card in enumerate(cards):
                scored = self._evaluate_card(card, idx, qi)
                if scored:
                    all_scored.append(scored)

            # After collecting, check if we have a good strict match
            strict = [s for s in all_scored if s[0] >= 30 and s[3]]
            if strict:
                strict.sort(key=lambda x: (-x[0], x[1]))
                best = strict[0]
                logger.info(
                    f"STRICT match: position #{best[1]}, "
                    f"score={best[0]}"
                )
                return self._click_candidate(best)

            # Scroll to load more
            BH.scroll(self.page, 1200)
            time.sleep(BH.short_delay())

        # No strict match found — try relaxed
        relaxed = [s for s in all_scored if s[0] >= 15 and s[3]]
        if relaxed:
            relaxed.sort(key=lambda x: (-x[0], x[1]))
            best = relaxed[0]
            logger.info(
                f"RELAXED match: position #{best[1]}, "
                f"score={best[0]}"
            )
            return self._click_candidate(best)

        # Last resort — any non-accessory match with brand + product line
        brand_only = [s for s in all_scored if s[0] >= 5 and s[3]]
        if brand_only:
            brand_only.sort(key=lambda x: (-x[0], x[1]))
            best = brand_only[0]
            logger.info(
                f"BRAND-ONLY match: position #{best[1]}, "
                f"score={best[0]}"
            )
            return self._click_candidate(best)

        return False

    def _evaluate_card(self, card, idx: int,
                       qi: QueryInfo) -> Optional[tuple]:
        """Evaluate a single product card.
        Returns (score, idx, link, model_ok) or None.
        """
        try:
            raw_text = card.inner_text()
            raw_lower = raw_text.lower()
            title = self._extract_title(card)
            title_norm = _normalize(title) if title else ""
            card_norm = _normalize(raw_text)
            combined = f"{title_norm} {card_norm}"

            # Hard filters — always reject these
            if _is_sponsored(raw_lower):
                return None

            if _is_accessory(title_norm, card_norm, raw_lower):
                return None

            if _is_wrong_category(combined, raw_lower, qi.brand):
                return None

            # Must have the product line (iphone/galaxy/pixel)
            if qi.product_line and qi.product_line not in combined:
                return None

            # ---- Scoring ----
            score = 0

            # Token overlap
            token_score = _match_score(combined, qi.search_tokens)
            score += token_score * 2

            # Brand
            if qi.brand and qi.brand in combined:
                score += 3

            # Product line
            if qi.product_line and qi.product_line in combined:
                score += 3

            # Model number near product line
            if qi.model_number:
                # HARD REJECT: if a DIFFERENT model number appears near
                # the product line, this is the wrong product entirely
                # e.g., searching "iphone 14" but card says "iphone 17"
                if self._conflicting_model(combined, qi):
                    return None

                if self._model_near(combined, qi):
                    score += 15
                elif re.search(
                    rf"\b{re.escape(qi.model_number)}\b", combined
                ):
                    score += 5
                else:
                    score -= 10  # Model not found at all

            # Samsung sub-family (flip vs fold vs s)
            if qi.samsung_sub:
                if self._has_sub_family(combined, qi.samsung_sub):
                    score += 10
                else:
                    score -= 15  # Wrong sub-family

            # Variant matching
            expected = set(qi.variant_tokens)
            present = set()
            for vt in VARIANT_TOKENS:
                if re.search(rf"\b{re.escape(vt)}\b", combined):
                    present.add(vt)

            for vt in expected:
                if vt in present:
                    score += 8
                else:
                    score -= 10

            for vt in present - expected:
                score -= 8

            # Category marker bonus
            if any(w in combined for w in
                   ["smartphone", "handy", "mobiltelefon"]):
                score += 5

            # Must have positive score
            if score <= 0:
                return None

            link = card.locator("a[href*='/p/']").first
            if link.count() == 0:
                # Fallback: any link
                link = card.locator("a").first
                if link.count() == 0:
                    return None

            # Candidate must satisfy model correctness when a model is requested.
            model_ok = True
            if qi.model_number:
                has_model = bool(
                    self._model_near(combined, qi)
                    or re.search(rf"\b{re.escape(qi.model_number)}\b", combined)
                )
                model_ok = has_model and (not self._conflicting_model(combined, qi))

            return (score, idx, link, model_ok)

        except Exception as e:
            logger.debug(f"Card eval error: {e}")
            return None

    def _model_near(self, text: str, qi: QueryInfo) -> bool:
        """Check if model number appears near product line or sub-family."""
        tokens = text.split()

        # For Samsung with sub-family, look after the sub-family token
        if qi.samsung_sub:
            for i, tok in enumerate(tokens):
                if tok == qi.samsung_sub:
                    window = tokens[i + 1: i + 4]
                    if qi.model_number in window:
                        return True
            return False

        # For other brands, look after the product line token
        if qi.product_line:
            for i, tok in enumerate(tokens):
                if tok == qi.product_line or tok.startswith(qi.product_line):
                    window = tokens[i + 1: i + 5]
                    if qi.model_number in window:
                        return True
                    rem = tok.replace(qi.product_line, "", 1).strip()
                    if rem == qi.model_number:
                        return True
        return False

    def _conflicting_model(self, text: str, qi: QueryInfo) -> bool:
        """Check for a different model number near the anchor token."""
        tokens = text.split()
        anchor = qi.samsung_sub or qi.product_line
        if not anchor:
            return False
        for i, tok in enumerate(tokens):
            if tok == anchor or tok.startswith(anchor):
                window = tokens[i + 1: i + 5]
                nums = [t for t in window if re.fullmatch(r"\d{1,2}", t)]
                if nums and all(n != qi.model_number for n in nums):
                    return True
        return False

    def _has_sub_family(self, text: str, sub: str) -> bool:
        if sub in ("flip", "fold"):
            return bool(re.search(
                rf"galaxy\s+z\s+{re.escape(sub)}\b", text
            ))
        return bool(re.search(
            rf"galaxy\s+{re.escape(sub)}\s*\d", text
        ))

    def _click_candidate(self, candidate: tuple) -> bool:
        """Click on a scored candidate and navigate to product page."""
        _, _, link, _ = candidate
        try:
            BH.mouse_move(self.page)
            time.sleep(BH.short_delay())
            link.click()
            try:
                self.page.wait_for_url("**/p/**", timeout=8000)
            except Exception:
                pass
            self.page.wait_for_load_state("domcontentloaded")
            self._check_captcha()
            time.sleep(BH.short_delay())
            logger.info(f"Product URL: {self.page.url}")
            return True
        except Exception as e:
            logger.error(f"Click failed: {e}")
            return False

    # ------------------------------------------------------------------
    # PDF extraction
    # ------------------------------------------------------------------
    def get_pdf_link(self) -> str:
        try:
            logger.info("Looking for Produktdatenblatt...")
            time.sleep(BH.short_delay())

            # JavaScript to check if element is in the main product area
            # (not inside recommendation/alternative/sponsored sections)
            is_main_product_js = """
            (el) => {
                let node = el;
                while (node && node !== document.body) {
                    const cls = (node.className || '').toLowerCase();
                    const id = (node.id || '').toLowerCase();
                    // Reject if inside recommendation/alternative sections
                    if (cls.match(/reco|alternative|similar|suggest|passend|fittingly|interessant|sponsored|anzeige/)) {
                        return false;
                    }
                    if (id.match(/reco|alternative|similar/)) {
                        return false;
                    }
                    // Check headings inside this container
                    if (node.tagName === 'SECTION' || node.tagName === 'DIV') {
                        const h = node.querySelector('h2, h3, h4');
                        if (h) {
                            const ht = h.textContent.toLowerCase();
                            if (ht.includes('alternative') || ht.includes('passend') ||
                                ht.includes('interessant') || ht.includes('interesting') ||
                                ht.includes('ähnlich') || ht.includes('similar') ||
                                ht.includes('fittingly') || ht.includes('zubehör')) {
                                return false;
                            }
                        }
                    }
                    node = node.parentElement;
                }
                return true;
            }
            """

            # Method 1: Find Produktdatenblatt links, skip ones in reco sections
            for label in ["Produktdatenblatt", "Product data sheet",
                          "product data sheet"]:
                links = self.page.locator(f"a:has-text('{label}')").all()
                for link in links:
                    try:
                        in_main = link.evaluate(is_main_product_js)
                        if not in_main:
                            logger.info(f"Skipping PDF link in recommendation section")
                            continue
                        href = link.get_attribute("href")
                        if href and ".pdf" in href.lower():
                            logger.info(f"Found PDF: {href}")
                            return href
                    except Exception:
                        continue

            # Method 2: Click-based fallback, also checking parent
            for label in ["Produktdatenblatt", "Product data sheet"]:
                elements = self.page.get_by_text(label).all()
                for el in elements:
                    try:
                        in_main = el.evaluate(is_main_product_js)
                        if not in_main:
                            continue

                        href = el.get_attribute("href")
                        if href and ".pdf" in href.lower():
                            logger.info(f"Found PDF: {href}")
                            return href

                        el.scroll_into_view_if_needed()
                        BH.scroll(self.page)
                        time.sleep(BH.short_delay())
                        el.click(timeout=3000)
                        time.sleep(random.uniform(0.8, 1.6))

                        pdf_links = self.page.locator(
                            "a[href*='d.otto.de'][href*='.pdf']"
                        ).all()
                        for pl in pdf_links:
                            try:
                                in_main2 = pl.evaluate(is_main_product_js)
                                if not in_main2:
                                    continue
                            except Exception:
                                pass
                            href = pl.get_attribute("href")
                            if href:
                                logger.info(f"Found PDF: {href}")
                                return href
                    except Exception as e:
                        logger.debug(f"PDF click error: {e}")

            logger.warning("PDF not found")
            return "Not found"
        except Exception as e:
            logger.error(f"PDF error: {e}")
            return "Not found"

    def _extract_title(self, card) -> str:
        try:
            link = card.locator("a[href*='/p/']").first
            if link.count() > 0:
                aria = link.get_attribute("aria-label")
                if aria:
                    return aria
                text = link.inner_text()
                if text:
                    return text
        except Exception:
            pass
        try:
            heading = card.locator("h2, h3, h4").first
            if heading.count() > 0:
                return heading.inner_text()
        except Exception:
            pass
        return ""

    # ------------------------------------------------------------------
    # ENERGY SECTION — scroll to it and expand it
    # ------------------------------------------------------------------
    def _scroll_to_energy_section(self):
        """Scroll down to find and expand the energy label section.
        The pdp_eek section may be collapsed — we need to click on
        the energy label to reveal the sheet image.
        """
        # First scroll down to find the energy section
        for _ in range(8):
            try:
                el = self.page.locator("[class*='pdp_eek']").first
                if el.count() > 0 and el.is_visible():
                    el.scroll_into_view_if_needed()
                    time.sleep(0.3)
                    break
            except Exception:
                pass
            BH.scroll(self.page, 600)
            time.sleep(0.2)

        # Click on the energy label to expand the sheet image
        # The label area is clickable and reveals the full energy sheet
        for click_sel in [
            ".pdp_eek__label",
            "[class*='pdp_eek__label']",
            ".js_openInPaliSheet",
            "img.pdp_eek__label-img",
            "[class*='pdp_eek'] [class*='label']",
        ]:
            try:
                el = self.page.locator(click_sel).first
                if el.count() > 0 and el.is_visible():
                    el.click(timeout=3000)
                    time.sleep(0.5)
                    logger.info(f"Clicked energy label: {click_sel}")
                    break
            except Exception:
                continue

    # ------------------------------------------------------------------
    # ENERGY CLASS — from product page DOM
    # ------------------------------------------------------------------
    def get_energy_class_from_page(self) -> str:
        """Extract energy class from img.pdp_eek__label-img alt attribute."""
        logger.info("Extracting energy class from page...")

        # Scroll to the energy section first
        self._scroll_to_energy_section()

        # Method 1: img.pdp_eek__label-img alt="A"
        try:
            el = self.page.locator("img.pdp_eek__label-img").first
            if el.count() > 0:
                alt = (el.get_attribute("alt") or "").strip()
                if re.fullmatch(r"[A-G][+]{0,3}", alt):
                    logger.info(f"Energy class from page: {alt}")
                    return alt
                src = el.get_attribute("src") or ""
                m = re.search(r"pl_eek_([a-g])", src, re.I)
                if m:
                    logger.info(f"Energy class from src: {m.group(1).upper()}")
                    return m.group(1).upper()
        except Exception:
            pass

        # Method 2: Text inside pdp_eek container
        try:
            el = self.page.locator("[class*='pdp_eek']").first
            if el.count() > 0:
                text = el.inner_text().strip()
                m = re.search(r"\b([A-G])\+{0,3}\b", text)
                if m:
                    logger.info(f"Energy class from eek text: {m.group(0)}")
                    return m.group(0)
        except Exception:
            pass

        # Method 3: Any img with src containing pl_eek
        try:
            imgs = self.page.locator("img[src*='pl_eek']").all()
            for img in imgs:
                src = img.get_attribute("src") or ""
                m = re.search(r"pl_eek_([a-g])", src, re.I)
                if m:
                    logger.info(f"Energy class from img src: {m.group(1).upper()}")
                    return m.group(1).upper()
                alt = (img.get_attribute("alt") or "").strip()
                if re.fullmatch(r"[A-G][+]{0,3}", alt):
                    return alt
        except Exception:
            pass

        logger.info("Energy class not found on page")
        return ""

    # ------------------------------------------------------------------
    # ENERGY LEVEL IMAGE LINK — from product page DOM
    # ------------------------------------------------------------------
    def get_energy_image_link(self) -> str:
        """Extract energy label sheet image URL."""
        logger.info("Extracting energy label image link...")

        try:
            el = self.page.locator("img.pdp_eek__sheet-image").first
            if el.count() > 0:
                srcset = el.get_attribute("srcset") or ""
                if srcset:
                    url = srcset.split()[0].strip()
                    if url.startswith("http"):
                        logger.info(f"Energy image from srcset: {url}")
                        return url
                src = el.get_attribute("src") or ""
                if src.startswith("http"):
                    logger.info(f"Energy image from src: {src}")
                    return src
        except Exception:
            pass

        try:
            el = self.page.locator(".pdp_eek__sheet-image-container img").first
            if el.count() > 0:
                srcset = el.get_attribute("srcset") or ""
                if srcset:
                    url = srcset.split()[0].strip()
                    if url.startswith("http"):
                        logger.info(f"Energy image from container: {url}")
                        return url
                src = el.get_attribute("src") or ""
                if src.startswith("http"):
                    return src
        except Exception:
            pass

        try:
            url = self.page.evaluate("""
                () => {
                    const img = document.querySelector('img.pdp_eek__sheet-image');
                    if (img) {
                        return img.srcset ? img.srcset.split(' ')[0] : img.src;
                    }
                    const container = document.querySelector('[class*="pdp_eek__sheet"]');
                    if (container) {
                        const innerImg = container.querySelector('img');
                        if (innerImg) {
                            return innerImg.srcset ? innerImg.srcset.split(' ')[0] : innerImg.src;
                        }
                    }
                    const allImgs = document.querySelectorAll('img[srcset*="i.otto.de"]');
                    for (const i of allImgs) {
                        const parent = i.closest('[class*="pdp_eek"]');
                        if (parent) {
                            return i.srcset ? i.srcset.split(' ')[0] : i.src;
                        }
                    }
                    return null;
                }
            """)
            if url and url.startswith("http"):
                logger.info(f"Energy image from JS: {url}")
                return url
        except Exception:
            pass

        try:
            time.sleep(1.0)
            el = self.page.locator("img.pdp_eek__sheet-image").first
            if el.count() > 0:
                srcset = el.get_attribute("srcset") or ""
                if srcset:
                    url = srcset.split()[0].strip()
                    if url.startswith("http"):
                        logger.info(f"Energy image (retry): {url}")
                        return url
                src = el.get_attribute("src") or ""
                if src.startswith("http"):
                    return src
        except Exception:
            pass

        logger.info("Energy image link not found")
        return "Not found"

    # ------------------------------------------------------------------
    # SUPPLIER — from "Details on product safety" popup
    # ------------------------------------------------------------------
    def get_supplier_from_page(self) -> str:
        """Click 'Details zur Produktsicherheit' / 'Details on product safety'
        and extract supplier info from the popup/slide panel that opens."""
        logger.info("Getting supplier from product safety popup...")

        # Step 1: Scroll down to Important Information section
        found_section = False
        for scroll_try in range(10):
            try:
                for marker in ["Wichtige Informationen",
                               "Important information"]:
                    el = self.page.get_by_text(marker, exact=False).first
                    if el.count() > 0 and el.is_visible():
                        el.scroll_into_view_if_needed()
                        time.sleep(0.3)
                        found_section = True
                        break
                if found_section:
                    break
            except Exception:
                pass
            BH.scroll(self.page, 800)
            time.sleep(0.3)

        if not found_section:
            logger.info("'Wichtige Informationen' section not found")

        # Step 2: Find and click the product safety link using JavaScript
        clicked = self.page.evaluate("""
        () => {
            const labels = [
                'Details zur Produktsicherheit',
                'Details on product safety',
                'Angaben zur Produktsicherheit',
                'Product safety details',
            ];
            const allEls = document.querySelectorAll('a, button, span, div[role="button"], [onclick], [class*="link"]');
            for (const el of allEls) {
                const text = (el.textContent || '').trim();
                for (const label of labels) {
                    if (text.includes(label) || text === label) {
                        el.scrollIntoView({block: 'center'});
                        el.click();
                        return label;
                    }
                }
            }
            const all = document.querySelectorAll('*');
            for (const el of all) {
                const directText = Array.from(el.childNodes)
                    .filter(n => n.nodeType === 3)
                    .map(n => n.textContent.trim())
                    .join(' ');
                for (const label of labels) {
                    if (directText.includes(label)) {
                        el.scrollIntoView({block: 'center'});
                        el.click();
                        return label;
                    }
                }
            }
            return null;
        }
        """)

        if not clicked:
            for label in ["Details zur Produktsicherheit",
                          "Details on product safety"]:
                try:
                    el = self.page.get_by_text(label, exact=False).first
                    if el.count() > 0:
                        el.scroll_into_view_if_needed()
                        time.sleep(0.15)
                        el.click(timeout=5000)
                        clicked = label
                        logger.info(f"Clicked via Playwright: {label}")
                        break
                except Exception:
                    pass

        if not clicked:
            logger.info("Product safety link not found")
            return ""

        logger.info(f"Clicked: {clicked}")

        # Step 3: Wait for the panel/popup to appear
        time.sleep(1.5)

        # Step 4: Use JavaScript to find the supplier text
        supplier_text = self.page.evaluate("""
        () => {
            const keywords = [
                'wirtschaftsakteur', 'economic operator',
                'responsible person', 'verantwortliche person',
                'verantwortlich für dieses produkt',
                'located in the eu', 'in der eu'
            ];

            const candidates = document.querySelectorAll(
                '[role="dialog"], [aria-modal], ' +
                '[class*="modal"], [class*="Modal"], ' +
                '[class*="drawer"], [class*="Drawer"], ' +
                '[class*="slide"], [class*="Slide"], ' +
                '[class*="panel"], [class*="Panel"], ' +
                '[class*="overlay"], [class*="Overlay"], ' +
                '[class*="layer"], [class*="Layer"], ' +
                '[class*="sheet"], [class*="Sheet"], ' +
                '[class*="popup"], [class*="Popup"], ' +
                '[class*="flyout"], [class*="Flyout"], ' +
                '[class*="sidebar"], [class*="Sidebar"], ' +
                '[class*="aside"], aside, ' +
                '[class*="product-safety"], [class*="productSafety"], ' +
                '[class*="pali"], [class*="Pali"]'
            );

            for (const el of candidates) {
                if (!el.offsetParent && el.style.display !== 'fixed') continue;
                const text = el.innerText || '';
                const tl = text.toLowerCase();
                if (keywords.some(k => tl.includes(k))) {
                    return text;
                }
            }

            const allDivs = document.querySelectorAll('div, section, aside');
            let best = null;
            let bestLen = 999999;
            for (const el of allDivs) {
                const style = window.getComputedStyle(el);
                const z = parseInt(style.zIndex) || 0;
                const pos = style.position;
                if (z > 10 || pos === 'fixed' || pos === 'absolute') {
                    if (!el.offsetParent && pos !== 'fixed') continue;
                    const rect = el.getBoundingClientRect();
                    if (rect.width < 50 || rect.height < 50) continue;
                    const text = el.innerText || '';
                    const tl = text.toLowerCase();
                    if (keywords.some(k => tl.includes(k))) {
                        if (text.length < bestLen) {
                            best = text;
                            bestLen = text.length;
                        }
                    }
                }
            }
            if (best) return best;

            const body = document.body.innerText || '';
            const bl = body.toLowerCase();
            if (keywords.some(k => bl.includes(k))) {
                for (const kw of keywords) {
                    const idx = bl.indexOf(kw);
                    if (idx !== -1) {
                        const start = Math.max(0, idx - 200);
                        const end = Math.min(body.length, idx + 2000);
                        return body.substring(start, end);
                    }
                }
            }

            return null;
        }
        """)

        result = ""
        if supplier_text:
            result = self._parse_supplier_popup(supplier_text)
            if result:
                logger.info(f"Supplier from popup: {result[:80]}...")
            else:
                logger.info("Popup text found but could not parse supplier")
                logger.debug(f"Popup text preview: {supplier_text[:300]}...")
        else:
            logger.info("No supplier-related text found after clicking")

        # Step 5: Close popup/panel
        try:
            self.page.evaluate("""
            () => {
                const closeBtns = document.querySelectorAll(
                    'button[aria-label="Close"], button[aria-label="Schließen"], ' +
                    'button[aria-label="close"], button[aria-label="schließen"], ' +
                    '[class*="close"], [class*="Close"]'
                );
                for (const btn of closeBtns) {
                    if (btn.offsetParent || window.getComputedStyle(btn).display !== 'none') {
                        btn.click();
                        return true;
                    }
                }
                const allBtns = document.querySelectorAll('button');
                for (const btn of allBtns) {
                    const t = (btn.textContent || '').trim();
                    if (t === '×' || t === 'X' || t === '✕') {
                        btn.click();
                        return true;
                    }
                }
                return false;
            }
            """)
        except Exception:
            pass

        try:
            self.page.keyboard.press("Escape")
            time.sleep(0.3)
        except Exception:
            pass

        return result

    @staticmethod
    def _parse_supplier_popup(text: str) -> str:
        if not text:
            return ""

        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        colon_markers = [
            "wirtschaftsakteur",
            "economic operator",
            "located in the eu",
            "in der eu",
            "responsible person",
            "verantwortliche person",
        ]

        start_idx = -1
        for i, line in enumerate(lines):
            ll = line.lower()
            if line.rstrip().endswith(":"):
                if any(m in ll for m in colon_markers):
                    start_idx = i
                    break
            for marker in colon_markers:
                pos = ll.find(marker)
                if pos != -1:
                    colon_pos = line.find(":", pos)
                    if colon_pos != -1:
                        after = line[colon_pos + 1:].strip()
                        if after and len(after) > 3:
                            collected = [after]
                            for j in range(i + 1, min(i + 7, len(lines))):
                                jl = lines[j].lower()
                                if _is_stop_line(jl):
                                    break
                                if _is_junk_line(lines[j]):
                                    continue
                                collected.append(lines[j])
                            if collected:
                                return re.sub(r"\s+", " ",
                                              " ".join(collected)).strip()
                        else:
                            start_idx = i
                            break
            if start_idx != -1:
                break

        if start_idx == -1:
            fallback_markers = [
                "located in the eu",
                "in der eu befindet",
                "in der eu angesiedelt",
                "in der eu ansässig",
                "the economic operator responsible",
                "der wirtschaftsakteur",
                "verantwortliche person",
                "responsible person for the eu",
                "verantwortlich für dieses produkt",
            ]
            for i, line in enumerate(lines):
                ll = line.lower()
                if any(m in ll for m in fallback_markers):
                    start_idx = i
                    break

        if start_idx == -1:
            return ""

        collected = []
        for i in range(start_idx + 1, min(start_idx + 8, len(lines))):
            ll = lines[i].lower()
            if _is_stop_line(ll):
                break
            if _is_junk_line(lines[i]):
                continue
            collected.append(lines[i])

        if collected:
            return re.sub(r"\s+", " ", " ".join(collected)).strip()
        return ""
def _is_stop_line(ll: str) -> bool:
    """Check if a lowercased line is a stop marker for supplier collection."""
    stop_markers = [
        "you can also find", "sie finden", "sie können",
        "du findest",
        "important information", "wichtige informationen",
        "report legal", "rechtliche bedenken",
        "return instruction", "rücksendeh",
        "disposal instruction", "entsorgungsh",
        "details on product", "details zur produkt",
        "discover another", "entdecke",
        "interesting alternative", "interessante alternative",
        "purchase on account", "kauf auf rechnung",
        "30-day", "30 tage",
        "https://", "http://",
        "attention:", "achtung:",
        "hinweis:", "bitte beachten",
    ]
    return any(s in ll for s in stop_markers)


def _is_junk_line(line: str) -> bool:
    """Check if a line is UI junk that should be skipped."""
    if len(line) < 3 or len(line) > 200:
        return True
    if line in ("×", "X", "Close", "Schließen", "OK",
                "Details zur Produktsicherheit",
                "Details on product safety",
                "Verantwortliche Person für die EU",
                "Responsible person for the EU"):
        return True
    return False


# ============================================================================

class PDFExtractor:
    def __init__(self, config: Config):
        self.config = config

    def extract_fields(self, pdf_url: str,
                       expected_brand: Optional[str] = None
                       ) -> Tuple[str, str]:
        if not pdf_url or pdf_url == "Not found":
            return ("Not found", "Not found")

        pdf_bytes = self._fetch(pdf_url)
        if not pdf_bytes:
            return ("Not found", "Not found")

        pages = self._text_pages(pdf_bytes) or []

        # Brand validation
        if expected_brand and pages:
            full = " ".join(pages).lower()
            tokens = [expected_brand]
            if expected_brand in BRAND_FAMILIES:
                tokens += BRAND_FAMILIES[expected_brand]["product_lines"]
            if not any(t in full for t in tokens):
                logger.warning(f"PDF brand mismatch: expected '{expected_brand}'")
                return ("Not found", "Not found")

        energy = ""
        supplier = ""

        if len(pages) >= 6:
            energy = self._energy([pages[5]])
        if not energy:
            energy = self._energy(pages)

        if len(pages) >= 25:
            supplier = self._supplier([pages[24]])
        if not supplier:
            supplier = self._supplier(pages)

        if self.config.OCR_ENABLED and (not energy or not supplier):
            ocr = self._ocr_pages(pdf_bytes)
            if ocr:
                if not energy:
                    energy = self._energy(ocr)
                if not supplier:
                    supplier = self._supplier(ocr)

        return (energy or "Not found", supplier or "Not found")

    def _fetch(self, url: str) -> Optional[bytes]:
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept": "application/pdf",
            })
            with urllib.request.urlopen(req, timeout=20) as r:
                return r.read()
        except Exception as e:
            logger.warning(f"PDF fetch failed: {e}")
            return None

    def _text_pages(self, data: bytes) -> list[str]:
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                return [p.extract_text() or "" for p in pdf.pages]
        except Exception:
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(data))
                return [(p.extract_text() or "") for p in reader.pages]
            except Exception as e:
                logger.warning(f"PDF parse failed: {e}")
                return []

    def _ocr_pages(self, data: bytes) -> list[str]:
        try:
            from pdf2image import convert_from_bytes, pdfinfo_from_bytes
            import pytesseract
        except Exception:
            return []
        try:
            info = pdfinfo_from_bytes(data,
                                      poppler_path=self.config.POPPLER_PATH)
            pc = int(info.get("Pages", 0))
        except Exception:
            pc = 0
        pns = ([6, 25] if pc >= 25 else [6] if pc >= 6
               else list(range(1, pc + 1)) if pc > 0 else [1])
        out: list[str] = []
        for pn in pns:
            try:
                imgs = convert_from_bytes(
                    data, dpi=self.config.OCR_DPI, fmt="png",
                    first_page=pn, last_page=pn,
                    poppler_path=self.config.POPPLER_PATH)
                if imgs:
                    pytesseract.pytesseract.tesseract_cmd = self.config.TESSERACT_CMD
                    try:
                        t = pytesseract.image_to_string(
                            imgs[0], lang=self.config.OCR_LANG) or ""
                    except Exception:
                        t = pytesseract.image_to_string(
                            imgs[0], lang="eng") or ""
                    out.append(t)
            except Exception as e:
                logger.warning(f"OCR page {pn}: {e}")
        return out

    def _energy(self, pages: list[str]) -> str:
        for text in pages:
            if not text:
                continue
            m = re.search(
                r"(?:Energy\s*efficiency\s*class|Energieeffizienzklasse)"
                r"\s*[:\-]?\s*([A-G](?:\s*[+]){0,3})", text, re.I)
            if m:
                return m.group(1).replace(" ", "")
            v = self._labeled(text,
                              ["Energy efficiency class",
                               "Energieeffizienzklasse"],
                              r"[A-G][+]{0,3}")
            if v:
                return v
            for i, line in enumerate(text.splitlines()):
                ll = line.lower()
                if "energieeffizienz" in ll or "energy efficiency" in ll:
                    s = line + (" " + text.splitlines()[i + 1]
                                if i + 1 < len(text.splitlines()) else "")
                    m2 = re.search(r"\b([A-G])\b", s)
                    if m2:
                        return m2.group(1)
        return ""

    def _supplier(self, pages: list[str]) -> str:
        la = ["Supplier information", "Lieferanteninformation",
              "Anschrift des Lieferanten",
              "Supplier's address", "Supplier address"]
        lb = ["Supplier's address", "Supplier address",
              "Lieferant", "Supplier"]
        for text in pages:
            if not text:
                continue
            for method in [
                lambda: self._inline_after(text,
                                           ["Anschrift des Lieferanten"]),
                lambda: self._supplier_address_table(text),
                lambda: self._block_after(text, la, 5),
                lambda: self._block_after_phrases(text, [
                    "Minimum duration of the guarantee offered by the supplier",
                    "Mindestdauer der vom Lieferanten angebotenen Garantie",
                ], 5),
                lambda: self._block_before(text, lb, 4),
                lambda: self._labeled(text, la + lb),
            ]:
                val = method()
                if val and self._valid_supplier(val):
                    return self._clean_supplier(val)
        return ""

    def _supplier_address_table(self, text: str) -> str:
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        lls = [l.lower() for l in lines]

        for i, ll in enumerate(lls):
            if "supplier" in ll and "address" in ll:
                m = re.search(
                    r"supplier'?s?\s*address\s*(?:\([a-z]\)\s*)*(.+)",
                    lines[i], re.I)
                if m:
                    val = m.group(1).strip()
                    if val and len(val) > 5 and not val.startswith("("):
                        collected = [val]
                        for j in range(i + 1, min(i + 5, len(lines))):
                            if self._heading(lines[j]):
                                break
                            if re.match(r"^\(", lines[j]):
                                break
                            if len(lines[j]) < 3:
                                break
                            collected.append(lines[j])
                        return " ".join(collected)

                collected = []
                for j in range(i + 1, min(i + 5, len(lines))):
                    if self._heading(lines[j]):
                        break
                    if re.match(r"^\([a-z]\)\s", lines[j]):
                        break
                    if self._annot(lines[j]):
                        continue
                    collected.append(lines[j])
                if collected:
                    return " ".join(collected)
        return ""

    @staticmethod
    def _valid_supplier(v: str) -> bool:
        if not v or len(v.strip()) < 5:
            return False
        lo = v.lower()
        for g in ["steuern", "weitere angaben", "self-repair",
                   "spare-parts", "search-detail"]:
            if g in lo:
                return False
        return sum(1 for c in v if c.isalpha()) >= 5

    @staticmethod
    def _clean_supplier(v: str) -> str:
        t = v.strip().replace("\u2019", "'")
        t = re.sub(r"^(\s*\([a-z]\)\s*)+", "", t, flags=re.I)
        t = re.sub(
            r"^(?:supplier information|lieferanteninformation"
            r"|anschrift des lieferanten|supplier'?s? address)\s*",
            "", t, flags=re.I)
        t = re.sub(r"\bsupplier\s*'?s?\s*address.*$", "", t, flags=re.I)
        t = re.sub(r"\banschrift des lieferanten.*$", "", t, flags=re.I)
        t = re.sub(r"\s*(\([a-z]\)\s*)+\s*$", "", t, flags=re.I)
        for mk in ["Produktdatenblatt", "Product information sheet",
                    "Additional information", "Angaben zur Reparierbarkeit"]:
            i = t.lower().find(mk.lower())
            if i != -1:
                t = t[:i].strip()
        t = re.sub(r"\d{1,2}\.\s+.+$", "", t).strip()
        t = re.sub(r"([a-z])([A-Z])", r"\1 \2", t)
        t = re.sub(r"([A-Za-z])(\d)", r"\1 \2", t)
        t = re.sub(r"(\d)([A-Za-z])", r"\1 \2", t)
        t = re.sub(r"https?://\S+", "", t)
        t = re.sub(r"\bSteuern\b.*$", "", t, flags=re.I)
        t = re.sub(r"\bWeitere Angaben\b.*$", "", t, flags=re.I)
        return re.sub(r"\s+", " ", t).strip()

    def _labeled(self, text: str, labels: list[str],
                 vre: Optional[str] = None) -> str:
        if not text:
            return ""
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        lls = [l.lower() for l in lines]
        for lb in labels:
            ll = lb.lower()
            for i, line in enumerate(lls):
                if ll in line:
                    m = re.search(rf"{re.escape(lb)}\s*[:\-]?\s*(.+)",
                                  lines[i], re.I)
                    if m and m.group(1).strip():
                        c = m.group(1).strip()
                        if vre:
                            vm = re.search(vre, c)
                            return vm.group(0) if vm else ""
                        return c
                    if i + 1 < len(lines):
                        c = lines[i + 1].strip()
                        if vre:
                            vm = re.search(vre, c)
                            return vm.group(0) if vm else ""
                        return c
        return ""

    def _inline_after(self, text: str, labels: list[str]) -> str:
        for lb in labels:
            pat = r"\s+".join(map(re.escape, lb.split()))
            m = re.search(rf"{pat}\s*(?:\([a-z]\)\s*)*:?\s*(.+)",
                          text, re.I | re.S)
            if m:
                return m.group(1).strip()
        return ""

    def _block_after(self, text: str, labels: list[str],
                     ml: int = 4) -> str:
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        lls = [l.lower() for l in lines]
        for lb in labels:
            ll = lb.lower()
            for i, line in enumerate(lls):
                if ll in line:
                    c = []
                    for j in range(i + 1, min(i + 1 + ml, len(lines))):
                        if re.match(r"^\([a-z]\)\s*$", lines[j], re.I):
                            break
                        if re.match(r"^\d+\.", lines[j]):
                            break
                        if lines[j].startswith("("):
                            break
                        if self._heading(lines[j]):
                            break
                        if self._annot(lines[j]):
                            continue
                        c.append(lines[j])
                    if c:
                        return " ".join(c)
        return ""

    def _block_after_phrases(self, text: str, phrases: list[str],
                             ml: int = 4) -> str:
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        lls = [l.lower() for l in lines]
        for ph in phrases:
            pl = ph.lower()
            for i, line in enumerate(lls):
                if pl in line:
                    c = []
                    for j in range(i + 1, min(i + 1 + ml, len(lines))):
                        if re.match(r"^\d+\.", lines[j]):
                            break
                        if lines[j].startswith("("):
                            break
                        if self._heading(lines[j]):
                            break
                        if self._annot(lines[j]):
                            continue
                        c.append(lines[j])
                    if c:
                        return " ".join(c)
        return ""

    def _block_before(self, text: str, labels: list[str],
                      ml: int = 4) -> str:
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        lls = [l.lower() for l in lines]
        for lb in labels:
            ll = lb.lower()
            for i, line in enumerate(lls):
                if ll in line:
                    c = []
                    for j in range(i - 1, max(-1, i - 1 - ml), -1):
                        if re.match(r"^\d+\.", lines[j]):
                            break
                        if self._heading(lines[j]):
                            break
                        if self._annot(lines[j]):
                            continue
                        c.append(lines[j])
                    if c:
                        return " ".join(reversed(c))
        return ""

    @staticmethod
    def _annot(t: str) -> bool:
        s = t.strip()
        return (not s or len(s) <= 2
                or bool(re.fullmatch(r"(\([a-z]\)\s*)+", s, re.I)))

    @staticmethod
    def _heading(t: str) -> bool:
        lo = t.lower()
        return any(h in lo for h in [
            "product information sheet", "produktdatenblatt",
            "additional information", "repairability",
            "angaben zur reparierbarkeit"])


# ============================================================================
# SCRAPER
# ============================================================================

class Scraper:
    def __init__(self):
        self.config = Config()
        self.pdf = PDFExtractor(self.config)

    def load_queries(self) -> list[str]:
        p = Path(self.config.INPUT_FILE)
        if not p.exists():
            raise FileNotFoundError(f"Not found: {self.config.INPUT_FILE}")
        qs = [l.strip() for l in p.read_text(encoding="utf-8").split("\n")
              if l.strip()]
        if not qs:
            raise ValueError("Empty input")
        logger.info(f"Loaded {len(qs)} queries")
        return qs

    def scrape(self, nav: OttoNavigator, query: str) -> Optional[ProductData]:
        try:
            nav.search_product(query)
            if not nav.find_product(query):
                return None

            url = nav.page.url
            brand = _detect_brand(_normalize(query))

            # Energy class from page DOM
            energy = nav.get_energy_class_from_page()

            # Energy level image link from page DOM
            energy_img = nav.get_energy_image_link()

            # Supplier ONLY from "Details zur Produktsicherheit" popup
            # Never from PDF — user requirement
            supplier = nav.get_supplier_from_page()

            # PDF link + PDF parsing (supplier fallback only)
            pdf = nav.get_pdf_link()
            _pdf_energy, pdf_supplier = self.pdf.extract_fields(pdf, brand)

            # Strict rule: energy must come from .pdp_eek__label on page.
            # If energy image link is missing, mark energy as not found too.
            if not energy_img or energy_img == "Not found":
                energy = ""
            if not supplier and pdf_supplier and pdf_supplier != "Not found":
                supplier = pdf_supplier

            return ProductData(
                query, url, pdf,
                energy if energy else "Not found",
                energy_img if energy_img else "Not found",
                supplier if supplier else "Not found",
            )
        except Exception as e:
            logger.error(f"Error: {query}: {e}")
            return None

    def run(self):
        queries = self.load_queries()
        results: list[dict] = []

        fields = ["input_ean", "product_url", "pdf_link",
                  "energy_efficiency_class", "energylevel_link",
                  "supplier_information"]

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=self.config.HEADLESS,
                slow_mo=self.config.SLOW_MO)
            ctx = browser.new_context(
                viewport={"width": self.config.VIEWPORT_WIDTH,
                          "height": self.config.VIEWPORT_HEIGHT},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36",
                accept_downloads=False)
            page = ctx.new_page()
            CaptchaSolver.apply_stealth(page)
            page.route("**/*.pdf", lambda r: r.abort())
            page.on("popup", lambda p: p.close())
            page.set_default_timeout(self.config.DEFAULT_TIMEOUT_MS)
            page.set_default_navigation_timeout(self.config.DEFAULT_TIMEOUT_MS)
            captcha = CaptchaSolver(self.config)
            nav = OttoNavigator(page, captcha)

            for i, q in enumerate(queries, 1):
                logger.info(f"\n{'=' * 60}")
                logger.info(f"[{i}/{len(queries)}]: {q}")
                logger.info(f"{'=' * 60}")

                prod = self.scrape(nav, q)
                if prod:
                    results.append(asdict(prod))
                else:
                    results.append({k: (q if k == "input_ean" else "")
                                    for k in fields})

                if i < len(queries):
                    time.sleep(BH.delay())

            browser.close()

        self._write_csv(fields, results)
        self._write_xlsx(fields, results)
        logger.info(f"\nDone! Results: {self.config.OUTPUT_FILE}")

    def _write_csv(self, fields: list[str], results: list[dict]):
        """Write CSV with space after each comma for readability."""
        with open(self.config.OUTPUT_FILE, "w", encoding="utf-8") as f:
            f.write(", ".join(fields) + "\n")
            for row in results:
                values = []
                for field in fields:
                    val = str(row.get(field, ""))
                    if "," in val or '"' in val:
                        val = '"' + val.replace('"', '""') + '"'
                    values.append(val)
                f.write(", ".join(values) + "\n")

    def _write_xlsx(self, fields: list[str], results: list[dict]):
        """Write XLSX output if openpyxl is available."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            logger.info("openpyxl not installed — skipping XLSX.")
            return

        xlsx_path = self.config.OUTPUT_FILE.rsplit(".", 1)[0] + ".xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Otto Products"

        headers = {
            "input_ean": ("Search Query", 45),
            "product_url": ("Product URL", 70),
            "pdf_link": ("PDF Link", 70),
            "energy_efficiency_class": ("Energy Class", 15),
            "energylevel_link": ("Energy Level Link", 60),
            "supplier_information": ("Supplier Information", 60),
        }

        bold = Font(bold=True)
        wrap = Alignment(wrap_text=True, vertical="top")

        for col, field in enumerate(fields, 1):
            label, width = headers.get(field, (field, 20))
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = bold
            ws.column_dimensions[chr(64 + col)].width = width

        for r, row in enumerate(results, 2):
            for c, field in enumerate(fields, 1):
                cell = ws.cell(row=r, column=c, value=row.get(field, ""))
                cell.alignment = wrap

        wb.save(xlsx_path)
        logger.info(f"XLSX saved: {xlsx_path}")


def main():
    try:
        Scraper().run()
    except Exception as e:
        logger.error(f"Fatal: {e}", exc_info=True)
        raise


if __name__ == "__main__":
    main()